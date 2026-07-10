#!/usr/bin/env -S uv run -s

# /// script
# requires-python = ">= 3.14"
# dependencies = [
#     "openpyxl",
#     "pydantic",
# ]
# ///

from datetime import datetime, date
from typing import cast

import openpyxl
import pydantic


class Event(pydantic.BaseModel):
    date_time: datetime = pydantic.Field(alias='date')
    details: str
    location: str
    url: str | None = None
    next: bool = False

    @property
    def date(self) -> date:
        return self.date_time.date()
    

class EventList(pydantic.BaseModel):
    events: list[Event]


def read_sheet() -> EventList:
    workbook = openpyxl.load_workbook("termine.xlsx")
    sheet = workbook.active
    if sheet is None:
        raise ValueError("No active sheet in the workbook")

    header = cast(
        list[str], list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    )

    events: list[Event] = []
    for row in sheet.iter_rows(min_row=2, max_col=4, values_only=True):
        event = Event(**dict(zip(header, row)))  # type: ignore
        events.append(event)

    return EventList(events=sorted(events, key=lambda x: x.date, reverse=True))


def main() -> None:
    print(read_sheet().model_dump_json())


if __name__ == "__main__":
    main()
